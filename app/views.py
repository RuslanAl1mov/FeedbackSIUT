import os
import requests

from django.shortcuts import render, get_object_or_404
from django.conf import settings

from .models import *

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import convertapi

from django.http import HttpResponse


def feedback_home_page(request):
    if request.method == "POST":
        user_language = request.POST.get('user_language')
        department_id = request.POST.get('department_id')
        subject_id = request.POST.get('subject_id')
        teacher_id = request.POST.get('teacher_id')

        teacher = get_object_or_404(Teacher, pk=teacher_id)
        subject = get_object_or_404(Subject, pk=subject_id)
        department = get_object_or_404(Department, pk=department_id)

        new_feedback = Feedback.objects.create(
            teacher=teacher,
            user_language=user_language,
            subject=subject,
            department=department
        )

        for i in range(9):
            Answer.objects.create(
                feedback=new_feedback,
                question_id=i + 1,
                answer_value=request.POST.get(f'question{i + 1}')
            )

    context = {"schools": School.objects.all()}
    return render(request, 'index.html', context=context)


def feedback_results_page(request):
    context = {'teachers': []}
    feedbacks = Feedback.objects.select_related('teacher', 'department', 'subject').all()

    existed_teachers = set()
    unique_teachers = [feedback.teacher for feedback in feedbacks if
                       feedback.teacher not in existed_teachers and not existed_teachers.add(feedback.teacher)]

    for teacher in unique_teachers:
        teacher_info = {'teacher': teacher, 'feedbacks': []}
        one_teacher_feedbacks = feedbacks.filter(teacher=teacher)

        existed_departments = set()
        unique_departments = [feedback.department for feedback in one_teacher_feedbacks if
                              feedback.department not in existed_departments and not existed_departments.add(
                                  feedback.department)]

        for department in unique_departments:
            department_feedbacks = {'department': department, 'subjects': [], 'feedbacks_number': 0}
            one_teacher_department_feedbacks = one_teacher_feedbacks.filter(department=department)

            existed_subjects = set()
            unique_subjects = [feedback.subject for feedback in one_teacher_department_feedbacks if
                               feedback.subject not in existed_subjects and not existed_subjects.add(feedback.subject)]

            for subject in unique_subjects:
                feedbacks_for_subject = one_teacher_department_feedbacks.filter(subject=subject)
                subject_feedbacks = {'subject': subject, 'answers': [], 'total_hor': []}

                for question_num in range(6):
                    answers = []
                    total_vert = []
                    for feedback in feedbacks_for_subject:
                        answer_obj = list(Answer.objects.filter(feedback=feedback).exclude(question_id__in=[7, 8, 9])
                                          .values('question_id', 'answer_value')
                                          .order_by('question_id'))

                        answers.append(answer_obj[question_num])
                        total_vert.append(int(answer_obj[question_num]['answer_value']))
                        if question_num == 0:
                            th = 0
                            for i in answer_obj:
                                th += int(i['answer_value'])
                            subject_feedbacks['total_hor'].append(format(th / len(answer_obj), '.2f'))
                    tv = 0
                    for i in total_vert:
                        tv += i
                    tv = tv / len(total_vert)

                    answers.append("{:.2f}".format(tv))
                    total_vert.clear()
                    subject_feedbacks['answers'].append(answers)

                th = 0
                for i in subject_feedbacks['total_hor']:
                    th += float(i)
                th = th / len(subject_feedbacks['total_hor'])
                subject_feedbacks['final_score'] = "{:.2f}".format(th)

                department_feedbacks['subjects'].append(subject_feedbacks)

            teacher_info['feedbacks'].append(department_feedbacks)

        context['teachers'].append(teacher_info)
        print(context)

    return render(request, 'feedback_results.html', context=context)


def download_person_feedback_report_OLD_VERSION(request, teacher_id):
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    feedbacks = Feedback.objects.select_related('department', 'subject').filter(teacher=teacher)

    context = {'teacher_info': teacher, 'feedbacks': []}

    existed_departments = set()
    unique_departments = [feedback.department for feedback in feedbacks if
                          feedback.department not in existed_departments and not existed_departments.add(
                              feedback.department)]

    for department in unique_departments:
        department_feedbacks = {'department': department, 'subjects': [], 'feedbacks_number': 0}
        department_feedbacks_for_teacher = feedbacks.filter(department=department)

        existed_subjects = set()
        unique_subjects = [feedback.subject for feedback in department_feedbacks_for_teacher if
                           feedback.subject not in existed_subjects and not existed_subjects.add(feedback.subject)]

        for subject in unique_subjects:
            feedbacks_for_subject = department_feedbacks_for_teacher.filter(subject=subject)
            subject_feedbacks = {'subject': subject, 'answers': [], 'total_hor': []}

            for question_num in range(6):  # Предполагается, что вопросов ровно 6
                answers = []
                total_vert = []
                for feedback in feedbacks_for_subject:
                    answer_obj = list(Answer.objects.filter(feedback=feedback).exclude(question_id__in=[7, 8, 9])
                                      .values('question_id', 'answer_value')
                                      .order_by('question_id'))

                    if answer_obj:  # Добавлено условие для проверки существования ответов
                        answers.append(answer_obj[question_num])
                        total_vert.append(int(answer_obj[question_num]['answer_value']))
                        if question_num == 0:
                            th = sum(int(i['answer_value']) for i in answer_obj)
                            subject_feedbacks['total_hor'].append(format(th / len(answer_obj), '.2f'))
                if total_vert:  # Проверка, чтобы избежать деления на ноль
                    tv = sum(total_vert) / len(total_vert)
                    answers.append("{:.2f}".format(tv))
                    total_vert.clear()
                    subject_feedbacks['answers'].append(answers)

            if subject_feedbacks['total_hor']:  # Проверка, чтобы избежать деления на ноль
                th = sum(float(i) for i in subject_feedbacks['total_hor']) / len(subject_feedbacks['total_hor'])
                subject_feedbacks['final_score'] = "{:.2f}".format(th)

            department_feedbacks['subjects'].append(subject_feedbacks)

        context['feedbacks'].append(department_feedbacks)

    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    thin_border = Border(left=Side(style='thin', color="DDDDDD"),
                         right=Side(style='thin', color="DDDDDD"),
                         top=Side(style='thin', color="DDDDDD"),
                         bottom=Side(style='thin', color="DDDDDD"))

    # Создание листа для каждого отделения
    for department_feedback in context['feedbacks']:
        ws = wb.create_sheet(title=department_feedback['department'].name[:31])  # Ограничиваем имя листа до 31 символа

        for c in range(20):
            column_letter = get_column_letter(c + 1)
            ws.column_dimensions[column_letter].width = 15

        for r in range(60):
            ws.row_dimensions[r].height = 20

        for subject in department_feedback['subjects']:
            sb_name_row = [subject['subject'].name]
            ws.append(sb_name_row)

            last_row = ws.max_row

            for col in range(1, len(sb_name_row) + 1):
                cell = ws.cell(row=last_row, column=col)
                cell.font = Font(bold=True, size=18)

            header_added = False
            question_num = 0
            for answers_list in subject['answers']:
                if not header_added:
                    table_headers = [f"Feedback {answer + 1}" for answer in range(len(answers_list))]
                    table_headers[-1] = "Total:"
                    table_headers.insert(0, '/')
                    ws.append([""])
                    ws.append(table_headers)
                    last_row = ws.max_row

                    for col in range(1, len(table_headers) + 1):
                        cell = ws.cell(row=last_row, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
                        cell.border = thin_border
                    header_added = True

                question_num += 1
                updated_an_list = [answer['answer_value'] for answer in answers_list if
                                   str(type(answer)) == "<class 'dict'>"]
                updated_an_list.append(answers_list[-1])
                updated_an_list.insert(0, f'Question {question_num}')
                ws.append(updated_an_list)
                last_row = ws.max_row

                for col in range(1, len(updated_an_list) + 1):
                    cell = ws.cell(row=last_row, column=col)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                cell = ws.cell(row=last_row, column=1)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
                cell.border = thin_border
                cell = ws.cell(row=last_row, column=len(updated_an_list))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
                cell.border = thin_border

            subject['total_hor'].insert(0, 'Average:')
            subject['total_hor'].append(subject['final_score'])
            ws.append(subject['total_hor'])
            last_row = ws.max_row

            for col in range(1, len(subject['total_hor']) + 1):
                cell = ws.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
                cell.border = thin_border

            ws.append([""])
            total_info = [f"Total: {subject['final_score']}"]
            ws.append(total_info)
            last_row = ws.max_row
            cell = ws.cell(row=last_row, column=1)
            cell.font = Font(bold=True, size=15)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(fill_type="solid", start_color="9FF5D7", end_color="9FF5D7")
            cell.border = thin_border

            ws.append([""])
            ws.append([""])
            ws.append([""])

    # Подготовка HttpResponse для скачивания файла
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{teacher.name}_feedback_report.xlsx"'
    wb.save(response)

    return response


def download_person_feedback_report(request, teacher_id):
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    feedbacks = Feedback.objects.select_related('department', 'subject').filter(teacher=teacher)

    context = {'teacher_info': teacher, 'feedbacks': []}

    existed_departments = set()
    unique_departments = [feedback.department for feedback in feedbacks if
                          feedback.department not in existed_departments and not existed_departments.add(
                              feedback.department)]

    for department in unique_departments:
        department_feedbacks = {'department': department, 'subjects': [], 'feedbacks_number': 0}
        department_feedbacks_for_teacher = feedbacks.filter(department=department)

        existed_subjects = set()
        unique_subjects = [feedback.subject for feedback in department_feedbacks_for_teacher if
                           feedback.subject not in existed_subjects and not existed_subjects.add(feedback.subject)]

        for subject in unique_subjects:
            feedbacks_for_subject = department_feedbacks_for_teacher.filter(subject=subject)
            subject_feedbacks = {'subject': subject, 'answers': [], 'total_hor': []}

            for question_num in range(6):  # Предполагается, что вопросов ровно 6
                answers = []
                total_vert = []
                for feedback in feedbacks_for_subject:
                    answer_obj = list(Answer.objects.filter(feedback=feedback).exclude(question_id__in=[7, 8, 9])
                                      .values('question_id', 'answer_value')
                                      .order_by('question_id'))

                    if answer_obj:  # Добавлено условие для проверки существования ответов
                        answers.append(answer_obj[question_num])
                        total_vert.append(int(answer_obj[question_num]['answer_value']))
                        if question_num == 0:
                            th = sum(int(i['answer_value']) for i in answer_obj)
                            subject_feedbacks['total_hor'].append(format(th / len(answer_obj), '.2f'))
                if total_vert:  # Проверка, чтобы избежать деления на ноль
                    tv = sum(total_vert) / len(total_vert)
                    answers.append("{:.2f}".format(tv))
                    total_vert.clear()
                    subject_feedbacks['answers'].append(answers)

            if subject_feedbacks['total_hor']:  # Проверка, чтобы избежать деления на ноль
                th = sum(float(i) for i in subject_feedbacks['total_hor']) / len(subject_feedbacks['total_hor'])
                subject_feedbacks['final_score'] = "{:.2f}".format(th)

            department_feedbacks['subjects'].append(subject_feedbacks)

        context['feedbacks'].append(department_feedbacks)

    # Построение пути к файлу
    file_path = os.path.join(settings.BASE_DIR, 'FeedbackSIUT/static/app/files/feedback_report_example.xlsx')
    excel_file_path = os.path.join(settings.BASE_DIR, 'FeedbackSIUT/static/app/files/cached.xlsx')
    pdf_file_path = os.path.join(settings.BASE_DIR, f'FeedbackSIUT/static/app/files/{teacher.name}_feedback_report.pdf')
    workbook = load_workbook(file_path)
    ws = workbook.active
    ws['B13'] = "Teacher's Name - " + teacher.name

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    i = 0
    total_of_total_score = 0
    for department in context['feedbacks']:
        for subject in enumerate(department['subjects']):
            i += 1
            cell_ref = f'B{i + 14}'
            ws[cell_ref] = i
            ws[cell_ref].border = thin_border
            cell_ref = f'C{i + 14}'
            ws[cell_ref] = department['department'].name
            ws[cell_ref].border = thin_border
            cell_ref = f'D{i + 14}'
            ws[cell_ref] = subject[1]['subject'].name
            ws[cell_ref].border = thin_border
            cell_ref = f'E{i + 14}'
            ws[cell_ref] = subject[1]['final_score']
            ws[cell_ref].border = thin_border
            ws[cell_ref].fill = PatternFill(fill_type="solid", start_color="CCFFCC", end_color="CCFFCC")
            total_of_total_score += float(subject[1]['final_score'])
    i += 1
    cell_ref = f'D{i + 14}'
    ws[cell_ref] = "Total:"
    ws[cell_ref].alignment = Alignment(horizontal="right")

    cell_ref = f'E{i + 14}'
    ws[cell_ref] = total_of_total_score

    workbook.save(filename=excel_file_path)
    workbook.close()

    # Ваш секретный ключ API от ConvertAPI
    convertapi.api_secret = os.getenv('CONVERTAPI_KEY')

    try:
        # Конвертация файла
        result = convertapi.convert('pdf', {
        'File': excel_file_path
        }, from_format='xlsx')
        result.file.save(pdf_file_path)

        # Отправляем PDF файл пользователю
        with open(pdf_file_path, 'rb') as pdf_file:
            django_response = HttpResponse(pdf_file.read(), content_type='application/pdf')
            django_response['Content-Disposition'] = f'attachment; filename="{os.path.basename(pdf_file_path)}"'
            return django_response
    except Exception as e:
        return HttpResponse(f"Ошибка при конвертации файла: {e}")
    finally:
        # Удаление PDF файла после отправки
        if os.path.exists(pdf_file_path):
            os.remove(pdf_file_path)
